import time
import requests
from rich.progress import track
from time import sleep
from colorama import Fore, init
from rich.console import Console
from rich.panel import Panel
import os
import sys
import datetime
from openpyxl import load_workbook, Workbook
from rich import text as RichText
import json
from pathlib import Path
import random

# --- Global Variables & Constants ---
init(autoreset=True)
console = Console()

Data_files = {
    'advice': 'data/advice_data.xlsx',
    'books': 'data/books_data.xlsx',
    'quotes': 'data/quotes_data.xlsx',
    'dogs': 'data/dogs_data.xlsx'
}

APIs = {
    'advice': 'https://api.adviceslip.com/advice',
    'books': 'https://potterapi-fedeperin.vercel.app/en/books',
    'quotes': 'https://hindi-quotes.vercel.app/random',
    'dogs': 'https://dog.ceo/api/breeds/image/random'
}

# --- UI Functions ---
def welcome():
    var = """
╔══════════════════════════════════════╗
║                                      ║
║         DAILY MOTIVATION HUB         ║
║  Your one-stop destination for daily ║
║             motivation!              ║
║  Books | Advice | Quotes | Cute Dogs ║
║                                      ║
╚══════════════════════════════════════╝"""
    print(Fore.GREEN + var)

def menu():
    var = """
╔════════════════ MAIN MENU ════════════════╗
║                                          ║
║   1. 💡 Get Random Advice                ║
║   2. 📚 Get Book Recommendation          ║
║   3. 🎯 Get Inspirational Quote            ║
║   4. 🐕 Get Cute Dog Image               ║
║   5. 🌟 Get Everything (Daily Dose!)     ║
║   6. 📊 Show Data Statistics             ║
║   7. ❌ Exit                             ║
║                                          ║
╚══════════════════════════════════════════╝"""
    print(Fore.BLUE + var)

def card(text, title, subtitle, color='Blue'):
    panel = Panel(text, style=color.lower(), title=title, subtitle=subtitle)
    console.print(panel)

# --- Setup and Utility Functions ---
def set_up_directories():
    '''Create necessary directories if they don't exist'''
    Path('data').mkdir(exist_ok=True)
    Path('log').mkdir(exist_ok=True)

def init_excel_files():
    """Initialize excel files with headers if they don't exist"""
    file_schemas = {
        'advice': ['Id', 'Advice', 'Timestamp'],
        'books': ['Title', 'Author', 'Description', 'Release Date', 'Pages', 'Timestamp'],
        'quotes': ['Quote', 'Author', 'Category', 'Timestamp'],
        'dogs': ['Image URL', 'Breed Info', 'Timestamp']
    }
    for file_type, headers in file_schemas.items():
        file_path = Data_files[file_type]
        if not os.path.exists(file_path):
            wb = Workbook()
            sheet = wb.active
            sheet.title = f'{file_type.capitalize()} Data'
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col, value=header)
            wb.save(file_path)

def show_loading():
    for _ in track(range(100), description='[green]Fetching data...'):
        sleep(0.01)

def time_now():
    return datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def log_error(error_message):
    """Log errors to a file"""
    with open('log/app.log', 'a') as log_file:
        log_file.write(f'[{time_now()}] ERROR: {error_message}\n')

# --- Core Logic ---
def make_api_request(api_name):
    """Make API request with error handling"""
    try:
        show_loading()
        response = requests.get(APIs[api_name], timeout=10)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        log_error(f'API error for {api_name}: {e}')
        return None
    except json.JSONDecodeError as e:
        log_error(f'JSON Decode Error for {api_name}: {e}')
        return None

def save_to_excel(data_type, data_row):
    """Save a single row of data to an excel file"""
    try:
        file_path = Data_files[data_type]
        wb = load_workbook(file_path)
        sheet = wb.active
        sheet.append(data_row)
        wb.save(file_path)
        return True
    except Exception as e:
        log_error(f'Excel Save Error for {data_type}: {e}')
        return False

# --- Feature Functions ---
def get_advice():
    """Fetch, display, and save random advice"""
    data = make_api_request('advice')
    if data and 'slip' in data:
        advice_id = data['slip']['id']
        advice_text = data['slip']['advice']
        card(title='💡 Advice for Today', text=f'"{advice_text}"', color='GREEN', subtitle=f'Advice ID: {advice_id}')
        if save_to_excel('advice', [advice_id, advice_text, time_now()]):
            print(f'{Fore.GREEN}✓ Advice saved successfully!')
        else:
            print(f'{Fore.RED}✗ Failed to save advice!')
    else:
        print(f'{Fore.RED}❌ Failed to fetch advice. Please try again later.')

def get_book():
    """Fetch, display, and save a random book recommendation"""
    data = make_api_request('books')
    if data:
        book = random.choice(data)
        book_info = {
            'Title': book.get('title', 'Unknown'),
            'Author': book.get('author', 'Unknown'),
            'Pages': book.get('pages', 'Unknown'),
            'Release Date': book.get('releaseDate', 'Unknown'),
            'Description': book.get('description', 'No description available.')
        }
        book_text = (
            f"[bold]Title:[/] {book_info['Title']}\n"
            f"[bold]Author:[/] {book_info['Author']}\n"
            f"[bold]Pages:[/] {book_info['Pages']}\n"
            f"[bold]Release Date:[/] {book_info['Release Date']}"
        )
        card(title='📚 Book Recommendation', text=book_text, color='CYAN', subtitle='Enjoy Reading..')
        if save_to_excel('books', list(book_info.values()) + [time_now()]):
            print(f'{Fore.GREEN}✓ Book recommendation saved to Excel!')
        else:
            print(f'{Fore.RED}✗ Failed to save book recommendation!')
    else:
        print(f'{Fore.RED}❌ Failed to fetch a book. Please try again later.')

def get_quote():
    """Fetch, display, and save an inspirational quote"""
    data = make_api_request('quotes')
    if data:
        quote_text = data.get('quote', 'No quote available.')
        author = data.get('author', 'Anonymous')
        category = data.get('type', 'General')
        card(title='🎯 Inspirational Quote', text=f'"{quote_text}"', color='YELLOW', subtitle=f"- {author}")
        if save_to_excel('quotes', [quote_text, author, category, time_now()]):
            print(f'{Fore.GREEN}✓ Quote saved to Excel!')
        else:
            print(f'{Fore.RED}✗ Failed to save quote.')
    else:
        print(f'{Fore.RED}❌ Failed to fetch a quote. Please try again later.')

def get_dog_image():
    """Fetch, display, and save a cute dog image URL"""
    data = make_api_request('dogs')
    if data and data.get('status') == 'success':
        image_url = data['message']
        breed_info = 'Random Dog'
        try:
            if "/breeds/" in image_url:
                breed_part = image_url.split('/breeds/')[1].split('/')[0]
                breed_info = ' '.join(word.capitalize() for word in breed_part.split('-')[::-1])
        except Exception:
            pass # Keep default breed_info if parsing fails
        dog_text = (
            f"[bold]Breed:[/] {breed_info}\n"
            f"[bold]Image URL:[/] {image_url}\n\n"
            f"Copy the URL into your browser to see the image!"
        )
        card(title='🐕 Cute Dog Image', text=dog_text, color='MAGENTA', subtitle='Here is a good boy/girl!')
        if save_to_excel('dogs', [image_url, breed_info, time_now()]):
            print(f'{Fore.GREEN}✓ Dog image info saved to Excel!')
        else:
            print(f'{Fore.RED}✗ Failed to save dog image info.')
    else:
        print(f'{Fore.RED}❌ Failed to fetch a dog image. Please try again later.')

def get_everything():
    """Fetch all content types for a daily dose of motivation"""
    print(f'\n{Fore.CYAN}🌟 Preparing your daily dose of motivation...\n')
    print(f"{Fore.YELLOW}{'='*60}")
    get_advice()
    print(f"{Fore.YELLOW}{'='*60}")
    get_book()
    print(f"{Fore.YELLOW}{'='*60}")
    get_quote()
    print(f"{Fore.YELLOW}{'='*60}")
    get_dog_image()
    print(f"\n{Fore.GREEN}Your daily motivation dose is complete! Have a great day. ✨")

def show_statistics():
    """Display statistics about saved data"""
    stats = {}
    total_entries = 0
    for data_type, file_path in Data_files.items():
        try:
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                sheet = wb.active
                count = sheet.max_row - 1
                stats[data_type] = max(0, count)
                total_entries += stats[data_type]
            else:
                stats[data_type] = 0
        except Exception:
            stats[data_type] = 0

    stats_text = (
        f"📝 [bold]Total Entries:[/] {total_entries}\n\n"
        f"💡 [bold]Advice Count:[/] {stats.get('advice', 0)}\n"
        f"📚 [bold]Books Count:[/] {stats.get('books', 0)}\n"
        f"🎯 [bold]Quotes Count:[/] {stats.get('quotes', 0)}\n"
        f"🐕 [bold]Dogs Count:[/] {stats.get('dogs', 0)}"
    )
    card(title='📊 Data Statistics', text=stats_text, color='CYAN', subtitle="Count of saved items")

# --- Main Application Runner ---
def main():
    """Main application function"""
    set_up_directories()
    init_excel_files()
    welcome()
    print('\n')

    while True:
        try:
            menu()
            choice = input(f"{Fore.CYAN}Enter your choice (1-7): ").strip()

            if choice == '1':
                get_advice()
            elif choice == '2':
                get_book()
            elif choice == '3':
                get_quote()
            elif choice == '4':
                get_dog_image()
            elif choice == '5':
                get_everything()
            elif choice == '6':
                show_statistics()
            elif choice == '7':
                print(f'{Fore.GREEN}Thank you for using Daily Motivation Hub! 🌟')
                print(f'{Fore.BLUE}Stay motivated and have a great day! 💪')
                break
            else:
                t = RichText("WARNING: Invalid choice. ", style="bold red")
                t.append("Only numbers from 1 to 7 are allowed!", style="italic yellow")
                console.print(t)
                time.sleep(2) # Pause briefly on invalid input

            if choice in ['1', '2', '3', '4', '5', '6']:
                input(f'\n{Fore.LIGHTBLACK_EX}Press Enter to continue...')
                print('\n' * 2)

        except KeyboardInterrupt:
            print(f'\n{Fore.YELLOW}👋 Goodbye! Thanks for using the Hub.')
            break
        except Exception as e:
            error_msg = f'An unexpected error occurred: {e}'
            print(f'{Fore.RED}❌ {error_msg}')
            log_error(f'Unexpected error in main loop: {e}')
            break

if __name__ == '__main__':
    main()